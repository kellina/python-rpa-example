from openpyxl import load_workbook
import time
import openpyxl
import pyautogui

pyautogui.PAUSE = 0.5

pyautogui.alert("O código vai começar. Não utilize nada do computador até o código finalizar!")

pyautogui.press('win')
pyautogui.write('libreoffice')
pyautogui.press('enter')
time.sleep(2)
pyautogui.click('word_icon.png')
time.sleep(2)
pyautogui.hotkey('win', 'up')
pyautogui.press('tab', presses= 3)
pyautogui.press('capslock')
pyautogui.write('top 5 produtos mais baratos', interval=0.2)
pyautogui.press('capslock')
pyautogui.press('enter', presses=3)
time.sleep(2)

file = 'C:\\Users\\114138\\Desktop\\planilha.xlsx'


def ler_data_xls(file: str):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    
    for value in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=2, values_only=True):
        pyautogui.write(value[0], interval=0.05)
        pyautogui.press('tab')
        pyautogui.write(str(value[1]))
        pyautogui.press('enter')
    
ler_data_xls(file)
pyautogui.hotkey('ctrl', 's')
pyautogui.write('lista de produtos', interval=0.1)
# pyautogui.click('area_trabalho.png')
# time.sleep(2)
# pyautogui.press('tab', presses=3)
time.sleep(1)
pyautogui.press('enter')
time.sleep(1)
pyautogui.hotkey('alt', 'f4')
pyautogui.press('left')
pyautogui.press('enter')


time.sleep(5)

pyautogui.press('win')
pyautogui.write('google')
pyautogui.press('enter')
time.sleep(2)
pyautogui.write('https://drive.google.com', interval=0.1)
pyautogui.press('enter')
# pyautogui.click('restjanela.png')
time.sleep(1)
pyautogui.hotkey('win','d')
pyautogui.moveTo(113,346)
pyautogui.mouseDown()
pyautogui.moveTo(1096, 532)

pyautogui.hotkey('alt','tab')
pyautogui.mouseUp()
time.sleep(5)
pyautogui.hotkey('alt', 'f4')
pyautogui.alert("O código foi finalizado. Você já pode utilizar o computador!")